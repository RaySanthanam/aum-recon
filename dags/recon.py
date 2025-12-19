from airflow import DAG
from airflow.sdk.definitions.decorators import task
from datetime import datetime, timedelta
from dbfread import DBF  # type: ignore
from pymongo import MongoClient
import sys
import os
import uuid
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from utils.sftp_client import SFTPClient

CAMS_DBF_PATH = "/opt/airflow/sftp_data/downloads/cams.dbf"
KARVEY_DBF_PATH = "/opt/airflow/sftp_data/downloads/karvey.dbf"
BATCH_SIZE = 2500
MONGO_URI = "mongodb://host.docker.internal:27017"
DB_NAME = "banking_demo"
TRANSACTIONS_COLLECTION = "wealth_pulse_transactions"
RECONCILIATION_COLLECTION = "reconciliation_results"
EXCEPTION_REPORT_PATH = "/opt/airflow/sftp_data/reports"


@task
def download_from_sftp():
    client = SFTPClient(
        host="sftp-server",
        port=22,
        username="testuser",
        password="testpass"
    )

    if client.connect():
        client.download_dbf_files()
        client.disconnect()


@task
def aggregate_mongodb_transactions():
    client = MongoClient("mongodb://host.docker.internal:27017")
    db = client["banking_demo"]
    collection = db["wealth_pulse_transactions"]

    aggregations_map = {}
    total_transactions = 0

    for txn in collection.find():
        folio = txn["folio_no"]
        scheme = txn["scheme_name"]
        product_code = txn["scheme_code"]
        qty = txn["units"]
        txn_type = txn["transaction_type"]

        key = f"{product_code}|{folio}|{scheme}"

        print(key)
        if key not in aggregations_map:
            aggregations_map[key] = 0.0

        if txn_type == "BUY":
            aggregations_map[key] += qty
        elif txn_type == "SELL":
            aggregations_map[key] -= qty

        total_transactions += 1

    for key in aggregations_map:
        aggregations_map[key] = round(aggregations_map[key], 3)

    client.close()

    print(f"✓ Processed {total_transactions} transactions")
    print(f"✓ Created {len(aggregations_map)} unique combinations")
    print("✓ SUCCESS: Aggregation completed!")
    return aggregations_map


@task
def reconcile_dbfs(aggregations_map):
    print(f"DEBUG: Received aggregations_map with {len(aggregations_map)} entries")

    matched = []
    mismatched = []
    dbf_only = []
    processed_mongo_keys = set()

    def process_dbf_record(key, dbf_units, source):
        if key in aggregations_map:
            if key in processed_mongo_keys:
                print(f"WARNING: Duplicate key found in {source}: {key}")
                return

            print(f"Matched Key from {source}: {key}")
            mongo_units = round(aggregations_map[key], 3)
            dbf_units_rounded = round(dbf_units, 3)

            record_data = {
                'source': source,
                'product_code': key[0],
                'folio': key[1],
                'scheme': key[2],
                'mongo_units': mongo_units,
                'dbf_units': dbf_units_rounded,
            }

            if abs(mongo_units - dbf_units_rounded) < 0.01:
                record_data['status'] = 'MATCHED'
                matched.append(record_data)
            else:
                record_data['difference'] = round(mongo_units - dbf_units_rounded, 3)
                record_data['status'] = 'MISMATCHED'
                mismatched.append(record_data)

            processed_mongo_keys.add(key)
        else:
            dbf_only.append({
                'source': source,
                'product_code': key[0],
                'folio': key[1],
                'scheme': key[2],
                'dbf_units': round(dbf_units, 3),
                'status': 'DBF_ONLY'
            })

    # Process CAMS DBF
    print("Processing CAMS DBF...")
    cams_table = DBF(CAMS_DBF_PATH, load=False)
    cams_count = 0

    for record in cams_table:
        product = record['PRODUCT'].strip()
        folio = record['FOLIO'].strip()
        scheme = record['SCHEME_NAM'].strip()
        dbf_units = float(record['UNITS'])
        key = f"{product}|{folio}|{scheme}"
        print(key)
        process_dbf_record(key, dbf_units, 'CAMS')
        cams_count += 1

    print(f"✓ Processed {cams_count} CAMS records")

    # Process Karvy DBF
    print("Processing Karvy DBF...")
    karvy_table = DBF(KARVEY_DBF_PATH, load=False)
    karvy_count = 0

    for record in karvy_table:
        product = record['PRCODE'].strip()
        folio = str(record['ACNO']).strip()
        scheme = record['FUNDDESC'].strip()
        dbf_units = float(record['BALUNITS'])
        key = f"{product}|{folio}|{scheme}"
        print(key)
        process_dbf_record(key, dbf_units, 'KARVY')
        karvy_count += 1

    print(f"✓ Processed {karvy_count} Karvy records")

    # Find MongoDB-only records
    mongo_only = []
    for key, units in aggregations_map.items():
        if key not in processed_mongo_keys:
            mongo_only.append({
                'product_code': key[0],
                'folio': key[1],
                'scheme': key[2],
                'mongo_units': round(units, 3),
                'status': 'MONGO_ONLY'
            })

    print("\n" + "=" * 70)
    print("RECONCILIATION SUMMARY")
    print("=" * 70)
    print(f"✓ Matched:        {len(matched)}")
    print(f"⚠ Mismatched:     {len(mismatched)}")
    print(f"📂 DBF Only:      {len(dbf_only)}")
    print(f"💾 MongoDB Only:  {len(mongo_only)}")
    print("=" * 70)

    return {
        'matched': matched,
        'mismatched': mismatched,
        'dbf_only': dbf_only,
        'mongo_only': mongo_only,
        'summary': {
            'matched_count': len(matched),
            'mismatched_count': len(mismatched),
            'dbf_only_count': len(dbf_only),
            'mongo_only_count': len(mongo_only),
        }
    }


@task
def store_reconciliation_results(reconciliation_data):
    client = MongoClient(MONGO_URI)
    db = client[DB_NAME]
    collection = db[RECONCILIATION_COLLECTION]

    collection.create_index([
        ("reconciliation_run_id", 1),
        ("status", 1)
    ])
    collection.create_index([
        ("folio", 1),
        ("product_code", 1),
        ("scheme", 1)
    ])
    collection.create_index([("reconciliation_date", -1)])
    collection.create_index([("status", 1)])

    run_id = str(uuid.uuid4())
    reconciliation_date = datetime.now()
    created_at = datetime.now()

    print(f"Reconciliation Run ID: {run_id}")
    print(f"Reconciliation Date: {reconciliation_date}")

    all_records = []

    # Store MATCHED records
    for record in reconciliation_data['matched']:
        doc = {
            'reconciliation_run_id': run_id,
            'reconciliation_date': reconciliation_date,
            'folio': record['folio'],
            'scheme': record['scheme'],
            'product_code': record['product_code'],
            'source': record['source'],
            'status': record['status'],
            'mongo_units': record['mongo_units'],
            'dbf_units': record['dbf_units'],
            'difference': 0.0,
            'created_at': created_at
        }
        all_records.append(doc)

    # Store MISMATCHED records
    for record in reconciliation_data['mismatched']:
        doc = {
            'reconciliation_run_id': run_id,
            'reconciliation_date': reconciliation_date,
            'folio': record['folio'],
            'scheme': record['scheme'],
            'product_code': record['product_code'],
            'source': record['source'],
            'status': record['status'],
            'mongo_units': record['mongo_units'],
            'dbf_units': record['dbf_units'],
            'difference': record['difference'],
            'created_at': created_at
        }
        all_records.append(doc)

    if all_records:
        result = collection.insert_many(all_records)
        inserted_count = len(result.inserted_ids)
        print(f"✓ Inserted {inserted_count} records into "
              f"{RECONCILIATION_COLLECTION}")
    else:
        inserted_count = 0
        print("No records to insert")

    client.close()

    print("\n" + "=" * 70)
    print("DATABASE STORAGE SUMMARY")
    print("=" * 70)
    print(f"Run ID:           {run_id}")
    print(f"Records Inserted: {inserted_count}")
    print(f"  - Matched:      {len(reconciliation_data['matched'])}")
    print(f"  - Mismatched:   {len(reconciliation_data['mismatched'])}")
    print(f"Collection:       {DB_NAME}.{RECONCILIATION_COLLECTION}")
    print("=" * 70)

    return {
        'run_id': run_id,
        'inserted_count': inserted_count,
        'reconciliation_date': reconciliation_date.isoformat()
    }


@task
def generate_exception_report(reconciliation_data, store_result):

    from pathlib import Path
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    Path(EXCEPTION_REPORT_PATH).mkdir(parents=True, exist_ok=True)

    run_id = store_result['run_id']
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    report_filename = f"exception_report_{timestamp}_{run_id[:8]}.xlsx"
    report_path = os.path.join(EXCEPTION_REPORT_PATH, report_filename)

    exception_records = []

    # Add MISMATCHED records
    for record in reconciliation_data['mismatched']:
        exception_records.append({
            'exception_type': 'MISMATCHED',
            'product_code': record['product_code'],
            'folio': record['folio'],
            'scheme': record['scheme'],
            'source': record['source'],
            'mongo_units': record.get('mongo_units', ''),
            'dbf_units': record.get('dbf_units', ''),
            'difference': record.get('difference', ''),
            'remarks': f"Unit mismatch: diff={record.get('difference', 0)}"
        })

    # Add DBF_ONLY records
    for record in reconciliation_data['dbf_only']:
        exception_records.append({
            'exception_type': 'DBF_ONLY',
            'product_code': record['product_code'],
            'folio': record['folio'],
            'scheme': record['scheme'],
            'source': record['source'],
            'mongo_units': '',
            'dbf_units': record.get('dbf_units', ''),
            'difference': '',
            'remarks': 'Record exists in DBF but not in MongoDB'
        })

    # Add MONGO_ONLY records
    for record in reconciliation_data['mongo_only']:
        exception_records.append({
            'exception_type': 'MONGO_ONLY',
            'product_code': record['product_code'],
            'folio': record['folio'],
            'scheme': record['scheme'],
            'source': 'MONGO',
            'mongo_units': record.get('mongo_units', ''),
            'dbf_units': '',
            'difference': '',
            'remarks': 'Record exists in MongoDB but not in any DBF file'
        })

    # Write Excel file
    if exception_records:
        wb = Workbook()
        ws = wb.active
        ws.title = "Exception Report"

        # Define headers
        headers = [
            'Exception Type', 'Product Code', 'Folio', 'Scheme',
            'Source', 'MongoDB Units', 'DBF Units', 'Difference', 'Remarks'
        ]

        # Style for header row
        header_fill = PatternFill(start_color="4472C4",
                                   end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers with styling
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write data rows
        for row_num, record in enumerate(exception_records, 2):
            ws.cell(row=row_num, column=1, value=record['exception_type'])
            ws.cell(row=row_num, column=2, value=record['product_code'])
            ws.cell(row=row_num, column=3, value=record['folio'])
            ws.cell(row=row_num, column=4, value=record['scheme'])
            ws.cell(row=row_num, column=5, value=record['source'])
            ws.cell(row=row_num, column=6, value=record['mongo_units'])
            ws.cell(row=row_num, column=7, value=record['dbf_units'])
            ws.cell(row=row_num, column=8, value=record['difference'])
            ws.cell(row=row_num, column=9, value=record['remarks'])

            # Color code exception types
            exception_type = record['exception_type']
            if exception_type == 'MISMATCHED':
                fill = PatternFill(start_color="FFF2CC",
                                   end_color="FFF2CC", fill_type="solid")
            elif exception_type == 'DBF_ONLY':
                fill = PatternFill(start_color="FCE4D6",
                                   end_color="FCE4D6", fill_type="solid")
            else:  # MONGO_ONLY
                fill = PatternFill(start_color="E2EFDA",
                                   end_color="E2EFDA", fill_type="solid")

            ws.cell(row=row_num, column=1).fill = fill

        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 50

        # Create Summary Sheet
        summary_ws = wb.create_sheet(title="Summary", index=0)

        # Summary title
        summary_ws['A1'] = 'RECONCILIATION SUMMARY'
        summary_ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        summary_ws['A1'].fill = PatternFill(start_color="203864",
                                            end_color="203864", fill_type="solid")
        summary_ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        summary_ws.merge_cells('A1:B1')

        # Run information
        summary_ws['A3'] = 'Run ID:'
        summary_ws['B3'] = run_id
        summary_ws['A4'] = 'Report Date:'
        summary_ws['B4'] = timestamp

        # Summary statistics
        summary_ws['A6'] = 'Category'
        summary_ws['B6'] = 'Count'

        # Style header
        for cell in ['A6', 'B6']:
            summary_ws[cell].font = Font(bold=True, color="FFFFFF")
            summary_ws[cell].fill = PatternFill(start_color="4472C4",
                                                end_color="4472C4", fill_type="solid")
            summary_ws[cell].alignment = Alignment(horizontal="center", vertical="center")

        # Total Matched (from reconciliation_data summary)
        total_matched = reconciliation_data['summary']['matched_count']
        summary_ws['A7'] = 'Total Matched'
        summary_ws['B7'] = total_matched
        summary_ws['A7'].fill = PatternFill(start_color="C6E0B4",
                                            end_color="C6E0B4", fill_type="solid")
        summary_ws['B7'].fill = PatternFill(start_color="C6E0B4",
                                            end_color="C6E0B4", fill_type="solid")

        # Total Mismatched
        total_mismatched = len(reconciliation_data['mismatched'])
        summary_ws['A8'] = 'Total Mismatched'
        summary_ws['B8'] = total_mismatched
        summary_ws['A8'].fill = PatternFill(start_color="FFF2CC",
                                            end_color="FFF2CC", fill_type="solid")
        summary_ws['B8'].fill = PatternFill(start_color="FFF2CC",
                                            end_color="FFF2CC", fill_type="solid")

        # Only in RTA (DBF_ONLY)
        total_rta_only = len(reconciliation_data['dbf_only'])
        summary_ws['A9'] = 'Only in RTA (DBF Files)'
        summary_ws['B9'] = total_rta_only
        summary_ws['A9'].fill = PatternFill(start_color="FCE4D6",
                                            end_color="FCE4D6", fill_type="solid")
        summary_ws['B9'].fill = PatternFill(start_color="FCE4D6",
                                            end_color="FCE4D6", fill_type="solid")

        # Only in Wealth Pulse (MONGO_ONLY)
        total_wealth_pulse_only = len(reconciliation_data['mongo_only'])
        summary_ws['A10'] = 'Only in Wealth Pulse (MongoDB)'
        summary_ws['B10'] = total_wealth_pulse_only
        summary_ws['A10'].fill = PatternFill(start_color="E2EFDA",
                                             end_color="E2EFDA", fill_type="solid")
        summary_ws['B10'].fill = PatternFill(start_color="E2EFDA",
                                             end_color="E2EFDA", fill_type="solid")

        # Total Exceptions
        summary_ws['A12'] = 'Total Exceptions'
        summary_ws['B12'] = len(exception_records)
        summary_ws['A12'].font = Font(bold=True)
        summary_ws['B12'].font = Font(bold=True)

        # Adjust column widths for summary sheet
        summary_ws.column_dimensions['A'].width = 30
        summary_ws.column_dimensions['B'].width = 20

        # Center align all values in column B
        for row in range(7, 14):
            summary_ws[f'B{row}'].alignment = Alignment(horizontal="center", vertical="center")

        # Save the workbook
        wb.save(report_path)

        print("\n" + "=" * 70)
        print("EXCEPTION REPORT GENERATED")
        print("=" * 70)
        print(f"Report File:      {report_filename}")
        print(f"Report Path:      {report_path}")
        print(f"Total Exceptions: {len(exception_records)}")
        print(f"  - Mismatched:   {len(reconciliation_data['mismatched'])}")
        print(f"  - DBF Only:     {len(reconciliation_data['dbf_only'])}")
        print(f"  - MongoDB Only: {len(reconciliation_data['mongo_only'])}")
        print("=" * 70)

        return {
            'report_path': report_path,
            'report_filename': report_filename,
            'total_exceptions': len(exception_records),
            'mismatched_count': len(reconciliation_data['mismatched']),
            'dbf_only_count': len(reconciliation_data['dbf_only']),
            'mongo_only_count': len(reconciliation_data['mongo_only'])
        }
    else:
        print("No exceptions found - all records matched!")
        return {
            'report_path': None,
            'total_exceptions': 0
        }


with DAG(
    dag_id="aum-recon-test",
    default_args={
        "owner": "batch_processing",
        "retries": 3,
        "retry_delay": timedelta(seconds=10),
    },
    start_date=datetime(2024, 1, 1),
    schedule=None,
    catchup=False,
) as dag:
    download_task = download_from_sftp()
    agg_map = aggregate_mongodb_transactions()
    recon_result = reconcile_dbfs(agg_map)
    store_task = store_reconciliation_results(recon_result)
    report_task = generate_exception_report(recon_result, store_task)

    download_task >> agg_map >> recon_result >> store_task >> report_task
